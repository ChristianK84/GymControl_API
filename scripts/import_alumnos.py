"""
Importador de alumnos desde formulario Excel (Google Forms) → BD GymControl.

Uso:
    python scripts/import_alumnos.py                  # dry-run por defecto
    python scripts/import_alumnos.py --commit          # escribe en BD
    python scripts/import_alumnos.py --archivo "ruta"  # otro archivo Excel

Requiere: openpyxl (pip install openpyxl si no está instalado)
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.core.database import SessionLocal
from app.models import Alumno, ContactoEmergencia, FichaMedica, Membresia, Transaccion, Tutor

MAESTRO_MAP = {
    "fernanda": 1,
    "natalia": 2,
    "grecia": 3,
    "oscar": 4,
    "jacky": 5,
    "jaqueline": 5,
    "martha": 6,
}

SKIP_VALUES = {"no", "ninguno", "ninguna", "n/a", "na", "none", "sin alergias", "sin condiciones", "sin", ""}


def normalizar_maestro(valor: str | None) -> int | None:
    if not valor:
        return None
    v = valor.strip().lower()
    v = v.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    v = v.replace("ü", "u").replace("ñ", "n")
    for key, mid in MAESTRO_MAP.items():
        if key in v:
            return mid
    return None


def limpiar_numero(valor) -> str | None:
    if valor is None:
        return None
    if isinstance(valor, float):
        return str(int(valor))
    s = str(valor).strip()
    return s if s else None


def parse_fecha(valor) -> date | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def limpiar_texto(valor) -> str | None:
    if valor is None:
        return None
    v = str(valor).strip()
    if not v or v.lower() in SKIP_VALUES | {"-"}:
        return None
    return v


def limpiar_nss(valor) -> str | None:
    """NSS: si es float → entero; si es texto tipo 'sí'/'No' → None."""
    if valor is None:
        return None
    if isinstance(valor, float):
        return str(int(valor))
    v = str(valor).strip()
    if not v or v.lower() in SKIP_VALUES | {"si", "sí"}:
        return None
    return v


def main():
    parser = argparse.ArgumentParser(description="Importar alumnos desde Excel a BD")
    parser.add_argument("--commit", action="store_true", help="Ejecutar escritura en BD")
    parser.add_argument(
        "--archivo",
        default=None,
        help="Ruta al archivo Excel (por defecto: Desktop/FormularioV3.xlsx)",
    )
    args = parser.parse_args()

    xlsx_path = args.archivo or (
        r"C:\Users\CRamos\OneDrive - Strattec Security Corporation\Desktop\FormularioV3.xlsx"
    )

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
    except Exception as e:
        print(f"ERROR: No se pudo abrir el archivo:\n  {e}")
        sys.exit(1)

    total = ws.max_row - 1  # fila 1 = headers

    print("=== GymControl - Importador de Alumnos (Excel -> BD) ===")
    print(f"Archivo:   {Path(xlsx_path).name}")
    print(f"Registros: {total}")
    print(f"Modo:      {'COMMIT' if args.commit else 'DRY-RUN (sin escritura)'}")
    print("-" * 72)

    db = SessionLocal() if args.commit else None

    if args.commit:
        print("Limpiando datos existentes...")
        db.execute(Transaccion.__table__.delete())
        db.execute(Alumno.__table__.delete())
        db.execute(Membresia.__table__.delete())
        db.commit()
        print("Datos anteriores eliminados.")

    stats = {"ok": 0, "skipped": 0, "errors": 0}

    for row_idx in range(2, ws.max_row + 1):
        vals = [cell.value for cell in ws[row_idx]]

        nombre = str(vals[1] or "").strip()
        apellido_p = str(vals[2] or "").strip()
        apellido_m = str(vals[3] or "").strip() or None
        fecha_nac = parse_fecha(vals[4])
        rama = str(vals[5] or "").strip()
        maestro_raw = str(vals[6] or "").strip() if vals[6] else None
        fotografia = str(vals[7] or "").strip() if vals[7] else None

        tutor_nombre = str(vals[8] or "").strip()
        tutor_ap_p = str(vals[9] or "").strip()
        tutor_ap_m = str(vals[10] or "").strip() or None
        tutor_tel = limpiar_numero(vals[11])
        tutor_email = str(vals[12] or "").strip() if vals[12] else None

        emergencia_nombre = str(vals[13] or "").strip()
        emergencia_ap_p = str(vals[14] or "").strip()
        emergencia_ap_m = str(vals[15] or "").strip() or None
        emergencia_tel = limpiar_numero(vals[16])

        tipo_sangre = str(vals[17] or "").strip() if vals[17] else None
        alergias = limpiar_texto(vals[18])
        medicamentos = limpiar_texto(vals[19])
        condiciones = limpiar_texto(vals[20])
        nss = limpiar_nss(vals[21])

        if not apellido_p and apellido_m:
            apellido_p = apellido_m
            apellido_m = None

        if tutor_ap_p == "-":
            tutor_ap_p = None
        if emergencia_ap_p == "-":
            emergencia_ap_p = None

        linea = f"Row {row_idx:>3} | {nombre} {apellido_p}"

        maestro_id = normalizar_maestro(maestro_raw) if maestro_raw else None

        motivos = []
        if not nombre:
            motivos.append("nombre vacio")
        if not apellido_p:
            motivos.append("apellido vacio")
        if not fecha_nac:
            motivos.append("fecha invalida")
        if not rama:
            motivos.append("rama vacia")
        if not maestro_id:
            motivos.append(f'maestro no reconocido: "{maestro_raw}"')

        if motivos:
            print(f"[SKIP] {linea} | Motivos: {'; '.join(motivos)}")
            stats["skipped"] += 1
            continue

        if args.commit:
            try:
                alumno = Alumno(
                    nombrecompleto=nombre,
                    apellido_paterno=apellido_p,
                    apellido_materno=apellido_m,
                    rama=rama,
                    fecha_nacimiento=fecha_nac,
                    maestro_id=maestro_id,
                    fotografia=fotografia,
                    fecha_inscripcion=date.today(),
                )
                db.add(alumno)
                db.flush()

                db.add(
                    Tutor(
                        alumno_id=alumno.id,
                        nombre=tutor_nombre,
                        apellido_paterno=tutor_ap_p,
                        apellido_materno=tutor_ap_m,
                        telefono=tutor_tel or "",
                        email=tutor_email or "",
                    )
                )

                db.add(
                    ContactoEmergencia(
                        alumno_id=alumno.id,
                        nombre=emergencia_nombre,
                        apellido_paterno=emergencia_ap_p,
                        apellido_materno=emergencia_ap_m,
                        telefono=emergencia_tel or "",
                    )
                )

                db.add(
                    FichaMedica(
                        alumno_id=alumno.id,
                        tipo_sangre=tipo_sangre,
                        alergias=alergias,
                        medicamentos=medicamentos,
                        condiciones_medicas=condiciones,
                        nss=nss,
                    )
                )

                db.commit()
            except Exception as e:
                db.rollback()
                print(f"[ERR]  {linea} | {e}")
                stats["errors"] += 1
                continue

        print(f"[OK]   {linea} | Maestro: ID={maestro_id} | Rama: {rama}")
        stats["ok"] += 1

    if args.commit:
        db.close()

    print("-" * 72)
    print("RESUMEN:")
    print(f"  OK:      {stats['ok']}")
    print(f"  Skipped: {stats['skipped']}")
    print(f"  Errores: {stats['errors']}")
    print("-" * 72)
    if not args.commit:
        print("Modo DRY-RUN: no se escribio nada en la BD.")
        print("Para importar: python scripts/import_alumnos.py --commit")
    else:
        print("Importacion completada.")


if __name__ == "__main__":
    main()
